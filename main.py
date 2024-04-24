import tkinter as tk
from tkinter import messagebox, filedialog
import psycopg2 #pip
import json
import os
import pandas as pd #pip
from openpyxl import load_workbook #pip
import random
import reportlab.lib #pip
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, Circle, String
from reportlab.lib.enums import TA_CENTER
import cv2 #pip
import pytesseract #pip and tesseract
import numpy as np
import platform


print(tk.TkVersion)
CONFIG_FILE = "config.json"
grade_scale = "grade_scale.json"

def save_config(host, username, password):
    config_data = {
        "host": host,
        "username": username,
        "password": password
    }
    with open(CONFIG_FILE, "w") as f:
        json.dump(config_data, f)

def load_config():
    try:
        with open(CONFIG_FILE, "r") as f:
            config_data = json.load(f)
            return config_data["host"], config_data["username"], config_data["password"]
    except FileNotFoundError:
        return "", "", ""

def delete_config():
    try:
        os.remove(CONFIG_FILE)
    except FileNotFoundError:
        pass

def login():
    host = host_entry.get()
    username = username_entry.get()
    password = password_entry.get()
    remember = remember_var.get()
    
    try:
        conn = psycopg2.connect(
            dbname="postgres",
            user=username,
            password=password,
            host=host,
            port="5432"
        )

        if remember:
            save_config(host, username, password)
        else:
            delete_config()
        
        root.destroy()
        print (conn.encoding)
        open_main_window(conn)  

    except psycopg2.Error as e:
        messagebox.showerror("Błąd logowania", "Błąd logowania do bazy danych:\n{}".format(e))
def get_subject_id(conn, subject_name):
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM subjects WHERE name = %s", (subject_name,))
        subject_id = cursor.fetchone()[0]
        cursor.close()
        return subject_id
    except psycopg2.Error as e:
        messagebox.showerror("Błąd", "Błąd podczas pobierania ID przedmiotu:\n{}".format(e))
        return None
    
def get_topic_id(conn, topic_name):
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM topics WHERE name = %s", (topic_name,))
        subject_id = cursor.fetchone()[0]
        cursor.close()
        return subject_id
    except psycopg2.Error as e:
        messagebox.showerror("Błąd", "Błąd podczas pobierania ID tematu:\n{}".format(e))
        return None
    


def open_main_window(conn):
    main_window = tk.Tk()
    main_window.title("Główne okno")
    main_window.geometry("600x450")

    subjects = get_subjects(conn)
    global topic_value

    subject_label = tk.Label(main_window, text="Wybierz przedmiot:")
    subject_label.grid(row=0, column=0)
    subject_value = tk.StringVar(main_window)
    subject_value.set(subjects[0]) if subjects else subject_value.set("Brak przedmiotów")

    subject_box = tk.OptionMenu(main_window, subject_value, ' ', *subjects)
    subject_box.grid(row=0, column=1)

    subject_plus_button = tk.Button(main_window, text="+", command=lambda: open_add_subject_window(conn, subject_value, subject_box))
    subject_plus_button.grid(row=0, column=2)
    
    subject_value.trace_add("write", lambda name, index, mode, sv=subject_value: on_subjectbox_select(conn, sv.get()))
    subject_value.trace_add("write", lambda *args: refresh_topicbox(conn, topic_box, subject_value.get()))
    topics = get_topics(conn, subject_value.get())
    
    
    
    
    topic_label = tk.Label(main_window, text="Wybierz temat:")
    topic_label.grid(row=1, column=0)
    topic_value = tk.StringVar(main_window)
    topic_value.set(topics[0]) if topics else topic_value.set("Brak tematów")

    topic_box = tk.OptionMenu(main_window, topic_value, ' ', *topics)
    topic_box.grid(row=1, column=1)

    topic_plus_button = tk.Button(main_window, text="+", command=lambda: open_add_topic_window(conn, topic_value, topic_box, subject_value.get()))
    topic_plus_button.grid(row=1, column=2)

    topic_value.trace_add("write", lambda name, index, mode, tv=topic_value: on_topicbox_select(conn, tv.get()))

    question_label = tk.Label(main_window, text="Dodaj pytanie: ")
    question_label.grid(row=2, column=0)
    question_plus_button = tk.Button(main_window, text="+", command=lambda: open_add_question_window(conn, topic_value.get()))
    question_plus_button.grid(row=2, column=1)
    question_import_button = tk.Button(main_window, text="Import z excela", command=lambda: open_import_question_window(conn, topic_value.get()))
    question_import_button.grid(row=2, column=2)

    generate_label = tk.Label(main_window, text="Wygeneruj test: ").grid(row=3, column=0)
    generate_button = tk.Button(main_window, text="Generuj", command=lambda: open_generate_window(conn, topic_value.get())).grid(row=3, column=1)

    check_label = tk.Label(main_window, text="Sprawdź test: ").grid(row=4, column=0)
    check_button = tk.Button(main_window, text="Sprawdź", command=lambda: open_opencv_window(conn)).grid(row=4, column=1)
    check_with_cam_button = tk.Button(main_window, text="Sprawdź za pomocą kamery", command=lambda: open_cam_check_window(conn)).grid(row=4, column=2)

    grades_button = tk.Button(main_window, text="Skala ocen", command=lambda: open_grading_scale(conn)).grid(row=5, column=0)


    main_window.mainloop()

def refresh_topicbox(conn, combo_box, current_subject):
    # Usunięcie istniejących elementów z comboboxa
    combo_box['menu'].delete(0, 'end')

    # Pobranie zaktualizowanej listy przedmiotów
    topics = get_topics(conn, current_subject)

    # Dodanie zaktualizowanej listy przedmiotów do comboboxa
    
    for topic in topics:
        combo_box['menu'].add_command(label=topic, command=lambda value=topic: topic_value.set(value))

def on_subjectbox_select(conn, selected_subject):
    subject_id = get_subject_id(conn, selected_subject)
    if subject_id is not None:
        print("Wybrano przedmiot o ID:", subject_id)
        return subject_id
    else:
        print("Nie można pobrać ID wybranego przedmiotu.")

def on_topicbox_select(conn, selected_topic):
    topic_id = get_topic_id(conn, selected_topic)
    if topic_id is not None:
        print("Wybrano temat o ID:", topic_id)
        return topic_id
    else:
        print("Nie można pobrać ID wybranego tematu.")

def open_add_subject_window(conn, subject_value, subject_box):
    add_window = tk.Toplevel()
    add_window.title("Dodaj przedmiot")
    add_window.geometry("200x100")

    tk.Label(add_window, text="Nazwa przedmiotu:").pack()
    subject_name_entry = tk.Entry(add_window)
    subject_name_entry.pack()

    add_button = tk.Button(add_window, text="Dodaj", command=lambda: add_subject(conn, subject_name_entry.get(), subject_value, subject_box, add_window))
    add_button.pack()

def open_add_topic_window(conn, topic_value, topic_box, current_subject):
    add_window = tk.Toplevel()
    add_window.title("Dodaj temat")
    add_window.geometry("200x100")

    tk.Label(add_window, text="Nazwa tematu:").pack()
    topic_name_entry = tk.Entry(add_window)
    topic_name_entry.pack()

    add_button = tk.Button(add_window, text="Dodaj", command=lambda: add_topic(conn, topic_name_entry.get(), topic_value, topic_box, add_window, current_subject))
    add_button.pack()

def open_add_question_window(conn, current_topic):
    add_window = tk.Toplevel()
    add_window.title("Dodaj pytanie")
    add_window.geometry("300x200")

    tk.Label(add_window, text="Pytanie:").grid(row=0, column=0)
    question_name_entry = tk.Entry(add_window)
    question_name_entry.grid(row=0, column=1)

    tk.Label(add_window, text="Odpowiedź poprawna:").grid(row=1, column=0)
    answer_1_entry = tk.Entry(add_window)
    answer_1_entry.grid(row=1, column=1)

    tk.Label(add_window, text="Odpowiedź 1:").grid(row=2, column=0)
    answer_2_entry = tk.Entry(add_window)
    answer_2_entry.grid(row=2, column=1)

    tk.Label(add_window, text="Odpowiedź 2:").grid(row=3, column=0)
    answer_3_entry = tk.Entry(add_window)
    answer_3_entry.grid(row=3, column=1)

    tk.Label(add_window, text="Odpowiedź 3:").grid(row=4, column=0)
    answer_4_entry = tk.Entry(add_window)
    answer_4_entry.grid(row=4, column=1)


    add_button = tk.Button(add_window, text="Dodaj", command=lambda: add_question(conn, question_name_entry.get(), answer_1_entry.get(), answer_2_entry.get(), answer_3_entry.get(), answer_4_entry.get(), add_window, current_topic))
    add_button.grid(row=5, column=0)

def open_import_question_window(conn, current_topic):
    filepath = filedialog.askopenfilename(title="Open Excel file", filetypes=[("Excel files", "*.xlsx")])
    
    if filepath:
        # Wczytanie danych z pliku Excela do ramki danych
        df = pd.read_excel(filepath)
        
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            # Tworzenie kursora
            cur = conn.cursor()
            current_topic = get_topic_id(conn, current_topic)

            for row in ws.iter_rows(min_row=2, values_only=True):

                data_to_insert = (current_topic,) + row
                insert_query = f"INSERT INTO qa (topic_id, question, correct_answer, answer_1, answer_2, answer_3) VALUES (%s, %s, %s, %s, %s, %s)"
                cur.execute(insert_query, data_to_insert)

            # Potwierdzenie transakcji
            conn.commit()

            # Zamykanie kursora i połączenia z bazą danych
            cur.close()

            print("Dane zostały pomyślnie dodane do bazy danych.")
        except psycopg2.Error as e:
            messagebox.showerror("Błąd", "Dane puste lub nieprawidłowe")
            print(f"Błąd: {e}")

def open_generate_window(conn, current_topic):
    generate_window = tk.Toplevel()
    generate_window.title("Generuj test")
    generate_window.geometry("300x200")

    tk.Label(generate_window, text="Nazwa pliku: ").grid(row=0, column=0)
    filename = tk.Entry(generate_window)
    filename.grid(row=0, column=1)

    tk.Label(generate_window, text="Ilość pytań: ").grid(row=1, column=0)
    question_quantity = tk.Entry(generate_window)
    question_quantity.grid(row=1, column=1)

    tk.Label(generate_window, text="Ilość grup: ").grid(row=2, column=0)
    group_quantity = tk.Entry(generate_window)
    group_quantity.grid(row=2, column=1)

    generate_button = tk.Button(generate_window, text="Generuj", command=lambda: generate_test(conn, current_topic, question_quantity.get(), filename.get(), group_quantity.get())).grid(row=3, column=0)
    print(question_quantity)

def open_opencv_window(conn):
    system = platform.system()
    filepath = filedialog.askopenfilename(title="Open Photo file", filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp")])
    image = cv2.imread(filepath)
    if system == "Windows":
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    elif system == "Darwin":
        pytesseract.pytesseract.tesseract_cmd = r"/usr/local/bin/tesseract"
    
    height, width, _ = image.shape
    
    # Określ proporcje obszaru zainteresowania (np. 10% szerokości i 5% wysokości)
    roi_width_percent = 15
    roi_height_percent = 20
    
    # Oblicz wymiary obszaru zainteresowania
    roi_width = int(width * roi_width_percent / 100)
    roi_height = int(height * roi_height_percent / 100)
    
    # Określ pozycję obszaru zainteresowania (w prawym górnym rogu obrazu)
    roi_x = width - roi_width
    roi_y = 0
    
    # Wycinamy obszar zainteresowania z obrazu
    roi = image[roi_y:roi_y+roi_height, roi_x:roi_x+roi_width]
    id = pytesseract.image_to_string(roi, config="--psm 7")
    print(id)
    
    
    image_height, image_width, _ = image.shape
    middle_region = image[:, image_width // 3:2 * image_width // 3]
    gray_middle = cv2.cvtColor(middle_region, cv2.COLOR_BGR2GRAY)

    # Wykonaj binaryzację obrazu dla środkowej części
    _, binary_middle = cv2.threshold(gray_middle, 150, 255, cv2.THRESH_BINARY_INV)

    # Znajdź kontury na obrazie binarnym dla środkowej części
    contours, _ = cv2.findContours(binary_middle, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Przefiltruj kontury, aby wykryć tylko te, które są odpowiednio dużymi prostokątami
    min_contour_area = 100
    max_contour_area = 1000
    filtered_contours = [cnt for cnt in contours if min_contour_area < cv2.contourArea(cnt) < max_contour_area]

    
    answers = {}
    for contour in filtered_contours:
        x, y, w, h = cv2.boundingRect(contour)
        roi = middle_region[y:y+h, x:x+w]
        mean_color = np.mean(roi)
        threshold = 160  # Prog wartości koloru, który oznacza, że obszar jest zamalowany
        print(mean_color)
        if mean_color < threshold:
            num_cols = 6  # Liczba obszarów kolumnowych na karcie odpowiedzi (4 odpowiedzi: A, B, C, D)
            col_width = middle_region.shape[1] // num_cols
            col_index = x // col_width
            labels = [' ', 'A', 'B', 'C', 'D', ' ']
            answers[(x, y)] = labels[col_index]
    # Wyświetl obraz z zaznaczonymi zamalowanymi prostokątami i etykietami odpowiedzi
    for (x, y), label in answers.items():
        cv2.putText(middle_region, label, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
    selected_answers = []
    sorted_answers = sorted(answers.items(), key=lambda item: item[0][1])
    for (_, _), label in sorted_answers:
        selected_answers.append(label)
    print (selected_answers)
    cv2.imshow('Detected Squares', middle_region)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
    check_test(conn, id, selected_answers)

def open_cam_check_window(conn):
    cap = cv2.VideoCapture(0)

    # Sprawdzenie czy kamera jest dostępna
    if not cap.isOpened():
        print("Błąd: Nie można otworzyć kamery.")
        return

    while True:
        # Przechwytywanie klatki z kamery
        ret, frame = cap.read()

        # Wywołanie funkcji do przetwarzania klatki
        system = platform.system()
    
        if system == "Windows":
            pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        elif system == "Darwin":
            pytesseract.pytesseract.tesseract_cmd = r"/usr/local/bin/tesseract"

        height, width, _ = frame.shape
        
        # Określ proporcje obszaru zainteresowania (np. 10% szerokości i 5% wysokości)
        roi_width_percent = 15
        roi_height_percent = 20
        
        # Oblicz wymiary obszaru zainteresowania
        roi_width = int(width * roi_width_percent / 100)
        roi_height = int(height * roi_height_percent / 100)
        
        # Określ pozycję obszaru zainteresowania (w prawym górnym rogu obrazu)
        roi_x = width - roi_width
        roi_y = 0
        
        # Wycinamy obszar zainteresowania z obrazu
        roi = frame[roi_y:roi_y+roi_height, roi_x:roi_x+roi_width]
        id = pytesseract.image_to_string(roi, config="--psm 7")
        print("ID:", id)
        
        image_height, image_width, _ = frame.shape
        middle_region = frame[:, image_width // 3:2 * image_width // 3]
        gray_middle = cv2.cvtColor(middle_region, cv2.COLOR_BGR2GRAY)

        # Wykonaj binaryzację obrazu dla środkowej części
        _, binary_middle = cv2.threshold(gray_middle, 150, 255, cv2.THRESH_BINARY_INV)

        # Znajdź kontury na obrazie binarnym dla środkowej części
        contours, _ = cv2.findContours(binary_middle, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # Przefiltruj kontury, aby wykryć tylko te, które są odpowiednio dużymi prostokątami
        min_contour_area = 200
        max_contour_area = 1000
        filtered_contours = [cnt for cnt in contours if min_contour_area < cv2.contourArea(cnt) < max_contour_area]

        answers = {}
        for contour in filtered_contours:
            x, y, w, h = cv2.boundingRect(contour)
            roi = middle_region[y:y+h, x:x+w]
            mean_color = np.mean(roi)
            threshold = 100  # Prog wartości koloru, który oznacza, że obszar jest zamalowany
            if mean_color < threshold:
                num_cols = 6  # Liczba obszarów kolumnowych na karcie odpowiedzi (4 odpowiedzi: A, B, C, D)
                col_width = middle_region.shape[1] // num_cols
                col_index = x // col_width
                labels = [' ', 'A', 'B', 'C', 'D', ' ']
                answers[(x, y)] = labels[col_index]
        
        # Wyświetl obraz z zaznaczonymi zamalowanymi prostokątami i etykietami odpowiedzi
        for (x, y), label in answers.items():
            cv2.putText(middle_region, label, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        
        selected_answers = []
        sorted_answers = sorted(answers.items(), key=lambda item: item[0][1])
        for (_, _), label in sorted_answers:
            selected_answers.append(label)
        print("Odpowiedzi:", selected_answers)
        cv2.imshow('Detected Squares', middle_region)

        # Wyjście z pętli po naciśnięciu klawisza 'q'
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    check_test(conn, id, selected_answers)  # Zakładając, że `conn` jest obiektem połączenia z bazą danych
    # Zwalnianie zasobów
    cap.release()
    cv2.destroyAllWindows()


def check_test(conn, id, selected_answers):
    id_test = id
    answ_from_db = []
    correct = 0
    cursor = conn.cursor()
    cursor.execute("SELECT a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20 FROM test WHERE id = %s", (id_test,))
    answ_from_db = cursor.fetchone()
    cursor.close()
    selected_answers = [answer.lower() for answer in selected_answers]
    answ_from_db_cleaned = []
    answ_from_db_cleaned = [answer for answer in answ_from_db if answer is not None]
    for i in range(len(selected_answers)):
        if selected_answers[i] == answ_from_db_cleaned[i]:
            correct += 1
    print(correct)
    correct_percent = (correct / len(selected_answers)) * 100
    print(correct_percent)
    with open("grade_scale.json", "r") as file:
        data = json.load(file)["grades"]
        if correct_percent >= data["grade1"]["grade_percent"]:
            grade_window = tk.Tk()
            grade_window.geometry("100x100")
            grade = tk.Label(grade_window, text=data["grade1"]["grade_value"], font=("Verdana", 40)).pack()
        elif correct_percent < data["grade1"]["grade_percent"] and correct_percent >= data["grade2"]["grade_percent"]:
            grade_window = tk.Tk()
            grade_window.geometry("100x100")
            grade = tk.Label(grade_window, text=data["grade2"]["grade_value"], font=("Verdana", 40)).pack()
        elif correct_percent < data["grade2"]["grade_percent"] and correct_percent >= data["grade3"]["grade_percent"]:
            grade_window = tk.Tk()
            grade_window.geometry("100x100")
            grade = tk.Label(grade_window, text=data["grade3"]["grade_value"], font=("Verdana", 40)).pack()
        elif correct_percent < data["grade3"]["grade_percent"] and correct_percent >= data["grade4"]["grade_percent"]:
            grade_window = tk.Tk()
            grade_window.geometry("100x100")
            grade = tk.Label(grade_window, text=data["grade4"]["grade_value"], font=("Verdana", 40)).pack()
        elif correct_percent < data["grade4"]["grade_percent"] and correct_percent >= data["grade5"]["grade_percent"]:
            grade_window = tk.Tk()
            grade_window.geometry("100x100")
            grade = tk.Label(grade_window, text=data["grade5"]["grade_value"], font=("Verdana", 40)).pack()
        elif correct_percent < data["grade5"]["grade_percent"]:
            grade_window = tk.Tk()
            grade_window.geometry("100x100")
            grade = tk.Label(grade_window, text=data["grade6"]["grade_value"], font=("Verdana", 40)).pack()

def generate_test(conn, current_topic, q_quantity, filename, group_quantity):
    for x in range(int(group_quantity)):
        if not q_quantity:
            messagebox.showerror("Błąd", "Wartość nie może być mniejsza lub równa 0")
            return
        topic_id = get_topic_id(conn, current_topic)
        cursor = conn.cursor()
        cursor.execute("SELECT question, correct_answer, answer_1, answer_2, answer_3 FROM qa WHERE topic_id = %s ORDER BY RANDOM() LIMIT %s", (topic_id, q_quantity))
        results = cursor.fetchall()
        cursor.close()

        qa = []
        ca = []
        for row in results:
            question, correct_answer, answer_1, answer_2, answer_3 = row
            answers = [correct_answer, answer_1, answer_2, answer_3]
            random.shuffle(answers)
            correct_answer_index = answers.index(correct_answer)
            ca.append((chr(97+correct_answer_index)))
            qa.append((question, answers))
        add_test(conn, ca)
        generate_pdf(qa, filename, current_topic, x)

def generate_pdf(questions_and_answers, filename, current_topic, current_group):
    filename += str(current_group+1) + ".pdf"
    doc = SimpleDocTemplate(filename, pagesize=letter)
    story = []

    title = current_topic + " Grupa: " + str(current_group + 1)

    pdfmetrics.registerFont(TTFont('Verdana', 'Verdana.ttf'))
    styles = getSampleStyleSheet()
    header_style = styles["Heading1"]
    body_style = styles["BodyText"]

    header_style.fontName = "Verdana"
    body_style.fontName = "Verdana"

    story.append(Paragraph(title, header_style))
    story.append(Spacer(1, 12))

    # Dodaj pytania i odpowiedzi do dokumentu PDF
    for index, (question, answers) in enumerate(questions_and_answers, start=1):
        question_text = f"Pytanie {index}: {question}"
        answers_text = "\n".join([f"{chr(97+i)}. {answer}" for i, answer in enumerate(answers)])
        
        story.append(Paragraph(question_text, body_style))
        story.append(Paragraph(answers_text, body_style))
        story.append(Spacer(1, 12))
    story.append(PageBreak())
    generate_answer_sheet(questions_and_answers, story, doc)
    doc.build(story)
    

def generate_answer_sheet(questions, story, doc):
    pdfmetrics.registerFont(TTFont('Symbol', 'Symbol.ttf'))
    styles = getSampleStyleSheet()
    body_style = styles["BodyText"]
    body_style.borderWidth = 1
    body_style.borderColor = "black"
    body_style.fontName = "Verdana"
    body_style.fontSize = 10
    body_style.align = "TA_CENTRE"
    top = [["Imię i nazwisko..................................................                     ", test_id]]

    table = Table(top, repeatRows=1)
    table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Wyśrodkuj zawartość pionowo
                               ('FONTNAME', (0, 0), (-1, -1), 'Verdana'),
                               ('FONTSIZE', (0, 0), (-1, -1), 16),
                               ('TEXTCOLOR', (0, 0), (-1, -1), colors.black)]))
    story.append(table)
    story.append(Spacer(1, 36))
    header = [["Nr zad.", "Odpowiedzi"]]
    table = Table(header, colWidths=[0.5*inch, 2*inch], repeatRows=1)
    table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Wyśrodkuj zawartość pionowo
                               ('FONTNAME', (0, 0), (-1, -1), 'Verdana'),
                               ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black),]))
    story.append(table)
    data = []
    for index, questions in enumerate(questions, start=1):
        answers = [Paragraph(f'\xa0\xa0{chr(97+i)}', body_style) for i in range(4)]
        data.append([index] + answers)

    table = Table(data, colWidths=[0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch], repeatRows=1)
    table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Wyśrodkuj zawartość pionowo
                               ('FONTNAME', (0, 0), (-1, -1), 'Verdana'),
                               ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                               ('GRID', (0, 0), (0, -1), 1, colors.black)]))
    table.spaceBefore = 10
    table.spaceAfter = 10
    
                               
    
    
    story.append(table)


def add_test(conn, ca):
    global test_id
    while len(ca) < 20:
        ca.append(None)
    cursor = conn.cursor()
    sql_query = ("INSERT INTO test (a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id")
    cursor.execute(sql_query, ca)
    test_id = cursor.fetchone()[0]
    conn.commit()

    cursor.close()
    print (test_id)
    return test_id


def add_subject(conn, subject_name, subject_value, subject_box, add_window):
    if not subject_name:
        messagebox.showerror("Błąd", "Nazwa przedmiotu nie może być pusta.")
        return
    
    try:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO subjects (name) VALUES (%s)", (subject_name,))
        conn.commit()
        cursor.close()
        messagebox.showinfo("Sukces", "Przedmiot dodany pomyślnie!")

        subjects = get_subjects(conn)
        subject_box['menu'].delete(0, 'end')  # Usunięcie poprzednich wartości
        for subject in subjects:
            subject_box['menu'].add_command(label=subject, command=tk._setit(subject_value, subject))

        add_window.destroy()

    except psycopg2.Error as e:
        messagebox.showerror("Błąd", "Błąd podczas dodawania przedmiotu:\n{}".format(e))

def add_topic(conn, topic_name, topic_value, topic_box, add_window, current_subject):
    if not topic_name:
        messagebox.showerror("Błąd", "Nazwa tematu nie może być pusta.")
        return
    
    try:
        cursor = conn.cursor()
        subject_id = get_subject_id(conn, current_subject)
        cursor.execute("INSERT INTO topics (name, subject_id) VALUES (%s, %s)", (topic_name, subject_id))
        conn.commit()
        cursor.close()
        messagebox.showinfo("Sukces", "Temat dodany pomyślnie!")

        topics = get_topics(conn, current_subject)
        topic_box['menu'].delete(0, 'end')  # Usunięcie poprzednich wartości
        for topic in topics:
            topic_box['menu'].add_command(label=topic, command=tk._setit(topic_value, topic))

        add_window.destroy()

    except psycopg2.Error as e:
        messagebox.showerror("Błąd", "Błąd podczas dodawania tematu:\n{}".format(e))

def add_question(conn, question_name, a1, a2, a3, a4, add_window, current_topic):
    if not question_name:
        messagebox.showerror("Błąd", "Nazwa pytania nie może być pusta.")
        return
    if not a1 and a2 and a3 and a4:
        messagebox.showerror("Błąd", "Odpowiedź nie może być pusta.")
        return
    
    try:
        cursor = conn.cursor()
        topic_id = get_topic_id(conn, current_topic)
        cursor.execute("INSERT INTO qa (topic_id, question, correct_answer, answer_1, answer_2, answer_3) VALUES (%s, %s, %s, %s, %s, %s)", (topic_id, question_name, a1, a2, a3, a4))
        conn.commit()
        cursor.close()
        messagebox.showinfo("Sukces", "Pytanie dodane pomyślnie!")

        add_window.destroy()

    except psycopg2.Error as e:
        messagebox.showerror("Błąd", "Błąd podczas dodawania pytania:\n{}".format(e))

def get_subjects(conn):
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM subjects")
        subjects = [row[0] for row in cursor.fetchall()]
        cursor.close()
        return subjects
    except psycopg2.Error as e:
        messagebox.showerror("Błąd", "Błąd podczas pobierania przedmiotów:\n{}".format(e))
        return []
    
def get_topics(conn, current_subject):
    try:
        subject_id = get_subject_id(conn, current_subject)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM topics WHERE subject_id = %s", (subject_id,))
        topics = [row[0] for row in cursor.fetchall()]
        cursor.close()
        return topics
    except psycopg2.Error as e:
        messagebox.showerror("Błąd", "Błąd podczas pobierania tematów:\n{}".format(e))
        return []

def open_grading_scale(conn):
    add_window = tk.Toplevel()
    add_window.title("Skala ocen")
    add_window.geometry("300x200")

    tk.Label(add_window, text="Ocena").grid(row=0, column=1)
    tk.Label(add_window, text="Procenty").grid(row=0, column=2)

    tk.Label(add_window, text="Ocena 1:").grid(row=1, column=0)
    grade1_entry = tk.Entry(add_window)
    grade1_entry.grid(row=1, column=1)
    grade1_percent = tk.Entry(add_window)
    grade1_percent.grid(row=1, column=2)

    tk.Label(add_window, text="Ocena 2:").grid(row=2, column=0)
    grade2_entry = tk.Entry(add_window)
    grade2_entry.grid(row=2, column=1)
    grade2_percent = tk.Entry(add_window)
    grade2_percent.grid(row=2, column=2)

    tk.Label(add_window, text="Ocena 3:").grid(row=3, column=0)
    grade3_entry = tk.Entry(add_window)
    grade3_entry.grid(row=3, column=1)
    grade3_percent = tk.Entry(add_window)
    grade3_percent.grid(row=3, column=2)

    tk.Label(add_window, text="Ocena 4:").grid(row=4, column=0)
    grade4_entry = tk.Entry(add_window)
    grade4_entry.grid(row=4, column=1)
    grade4_percent = tk.Entry(add_window)
    grade4_percent.grid(row=4, column=2)

    tk.Label(add_window, text="Ocena 5:").grid(row=5, column=0)
    grade5_entry = tk.Entry(add_window)
    grade5_entry.grid(row=5, column=1)
    grade5_percent = tk.Entry(add_window)
    grade5_percent.grid(row=5, column=2)

    tk.Label(add_window, text="Ocena 6:").grid(row=6, column=0)
    grade6_entry = tk.Entry(add_window)
    grade6_entry.grid(row=6, column=1)

    load_grade_data(grade1_entry, grade1_percent,
                    grade2_entry, grade2_percent,
                    grade3_entry, grade3_percent,
                    grade4_entry, grade4_percent,
                    grade5_entry, grade5_percent,
                    grade6_entry)

    save_button = tk.Button(add_window, text="Zapisz", command=lambda: save_grade_scale(grade1_entry.get(), grade1_percent.get(),
                                                                                        grade2_entry.get(), grade2_percent.get(),
                                                                                        grade3_entry.get(), grade3_percent.get(),
                                                                                        grade4_entry.get(), grade4_percent.get(),
                                                                                        grade5_entry.get(), grade5_percent.get(),
                                                                                        grade6_entry.get()))
    save_button.grid(row=7, column=0)
    
def save_grade_scale(grade1_value, grade1_percent,
                    grade2_value, grade2_percent,
                    grade3_value, grade3_percent,
                    grade4_value, grade4_percent,
                    grade5_value, grade5_percent,
                    grade6_value):
    grades_data = {
        "grade1": {
            "grade_value": grade1_value,
            "grade_percent": float(grade1_percent)
        },
        "grade2": {
            "grade_value": grade2_value,
            "grade_percent": float(grade2_percent)
        },
        "grade3": {
            "grade_value": grade3_value,
            "grade_percent": float(grade3_percent)
        },
        "grade4": {
            "grade_value": grade4_value,
            "grade_percent": float(grade4_percent)
        },
        "grade5": {
            "grade_value": grade5_value,
            "grade_percent": float(grade5_percent)
        },
        "grade6": {
            "grade_value": grade6_value
        }

    }
    
    save_scale = {
        "grades": grades_data
    }
    with open(grade_scale, "w") as f:
        json.dump(save_scale, f)

def load_grade_data(
        grade1_entry, grade1_percent,
        grade2_entry, grade2_percent,
        grade3_entry, grade3_percent,
        grade4_entry, grade4_percent,
        grade5_entry, grade5_percent,
        grade6_entry
    ):
    try:
        with open("grade_scale.json", "r") as file:
            data = json.load(file)["grades"]
            grade1_entry.insert(0, data["grade1"]["grade_value"])
            grade1_percent.insert(0, data["grade1"]["grade_percent"])
            grade2_entry.insert(0, data["grade2"]["grade_value"])
            grade2_percent.insert(0, data["grade2"]["grade_percent"])
            grade3_entry.insert(0, data["grade3"]["grade_value"])
            grade3_percent.insert(0, data["grade3"]["grade_percent"])
            grade4_entry.insert(0, data["grade4"]["grade_value"])
            grade4_percent.insert(0, data["grade4"]["grade_percent"])
            grade5_entry.insert(0, data["grade5"]["grade_value"])
            grade5_percent.insert(0, data["grade5"]["grade_percent"])
            grade6_entry.insert(0, data["grade6"]["grade_value"])
    except FileNotFoundError:
        pass

root = tk.Tk()
root.title("Logowanie do PGAdmina")
root.geometry("400x300")

host, username, password = load_config()

tk.Label(root, text="Host:").pack()
host_entry = tk.Entry(root)
host_entry.pack()
host_entry.insert(0, host)

tk.Label(root, text="Login:").pack()
username_entry = tk.Entry(root)
username_entry.pack()
username_entry.insert(0, username)

tk.Label(root, text="Hasło:").pack()
password_entry = tk.Entry(root, show="*")
password_entry.pack()
password_entry.insert(0, password)

remember_var = tk.BooleanVar()
remember_checkbox = tk.Checkbutton(root, text="Zapamiętaj dane logowania", variable=remember_var)
remember_checkbox.pack()
remember_checkbox.select()

login_button = tk.Button(root, text="Zaloguj", command=login)
login_button.pack()

root.mainloop()
